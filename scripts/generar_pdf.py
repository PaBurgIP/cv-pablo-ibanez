#!/usr/bin/env python3
"""Genera un CV documental en PDF a partir del HTML manual y Actividades.xlsx.

La salida no imprime la web: compone un documento A4 resumido, pensado para
lectura, archivo y envío profesional.
"""

from __future__ import annotations

import argparse
import html
import os
import re
import tempfile
from datetime import date, datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    HRFlowable,
    KeepTogether,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from traducir_cv import translate_value


CV_DIR = Path(__file__).resolve().parents[1]
EXCEL_DIR = CV_DIR.parent / "Certificados Académicos"
EXCEL_CANDIDATES = [EXCEL_DIR / "Actividades.xlsx", EXCEL_DIR / "Actividades.actualizado.xlsx"]
DEFAULT_EXCEL = max(
    (candidate for candidate in EXCEL_CANDIDATES if candidate.exists()),
    key=lambda candidate: candidate.stat().st_mtime,
    default=EXCEL_CANDIDATES[0],
)

GREEN = colors.HexColor("#2F8F52")
DARK = colors.HexColor("#24343B")
MUTED = colors.HexColor("#5D6A70")
CLAY = colors.HexColor("#B5651D")
LINE = colors.HexColor("#D7DFDA")
PALE = colors.HexColor("#F2F6F2")

LABELS = {
    "es": {
        "title": "Investigador predoctoral en CISA-INIA-CSIC",
        "profile": "PERFIL PROFESIONAL",
        "profile_text": (
            "Investigador predoctoral especializado en epidemiología espacial, modelización matemática, "
            "ciencia de datos y sistemas de información geográfica aplicados a vigilancia y evaluación "
            "de riesgos bajo el enfoque One Health."
        ),
        "experience": "EXPERIENCIA EN PROYECTOS DE INVESTIGACIÓN",
        "education": "FORMACIÓN ACADÉMICA",
        "publications": "PUBLICACIONES CIENTÍFICAS",
        "technical": "ARTÍCULOS TÉCNICO-PROFESIONALES Y DE TRANSFERENCIA",
        "activity": "ACTIVIDAD CIENTÍFICA",
        "activity_summary": "{forums} encuentros científicos registrados · {lead} contribuciones como autor principal o ponente · {international} contribuciones internacionales.",
        "lead_contributions": "Contribuciones como autor principal",
        "teaching": "DOCENCIA, ASESORÍA Y TRANSFERENCIA",
        "courses": "FORMACIÓN CONTINUA SELECCIONADA",
        "courses_summary": "Selección de cursos dentro de {hours} horas acreditadas de formación continua.",
        "tools": "HERRAMIENTAS DESARROLLADAS",
        "outreach": "DIVULGACIÓN CIENTÍFICA",
        "recognition": "PREMIOS, BECAS Y ESTANCIAS",
        "skills": "COMPETENCIAS",
        "advice": "Asesoría técnica sobre la situación de la Influenza Aviar de Alta Patogenicidad (IAAP) y el uso de DiFLUsion como sistema de alerta temprana",
        "page": "Página",
    },
    "en": {
        "title": "Predoctoral Researcher at CISA-INIA-CSIC",
        "profile": "PROFESSIONAL PROFILE",
        "profile_text": (
            "Predoctoral researcher specialising in spatial epidemiology, mathematical modelling, data "
            "science and geographic information systems for One Health surveillance and risk assessment."
        ),
        "experience": "RESEARCH PROJECT EXPERIENCE",
        "education": "EDUCATION",
        "publications": "SCIENTIFIC PUBLICATIONS",
        "technical": "TECHNICAL AND KNOWLEDGE-TRANSFER ARTICLES",
        "activity": "SCIENTIFIC ACTIVITY",
        "activity_summary": "{forums} scientific events recorded · {lead} lead-author or speaker contributions · {international} international contributions.",
        "lead_contributions": "Lead-author contributions",
        "teaching": "TEACHING, ADVICE AND KNOWLEDGE TRANSFER",
        "courses": "SELECTED CONTINUING EDUCATION",
        "courses_summary": "Selected courses from {hours} accredited hours of continuing education.",
        "tools": "DEVELOPED TOOLS",
        "outreach": "SCIENCE OUTREACH",
        "recognition": "AWARDS, FELLOWSHIPS AND RESEARCH STAYS",
        "skills": "SKILLS",
        "advice": "Technical advice on the highly pathogenic avian influenza (HPAI) situation and the use of DiFLUsion as an early-warning system",
        "page": "Page",
    },
}

COURSE_SELECTION = (
    "Análisis de la biodiversidad de especies y hábitats con QGIS, Google Earth Engine y OpenData",
    "Digitalización: Agricultura 4.0",
    "Python avanzado",
    "Scripts de geoprocesamiento con Python en ArcGIS Pro",
    "Machine Learning para Investigación Científica - Random Forest, Boosting y Técnicas Avanzadas de Interpretación",
    "Deep learning con KERAS para investigación científica",
    "Desarrollo web con ArcGIS API para JavaScript",
    "Servicios sobre Docker",
    "Adapted numerical methods for Differential Equations",
    "FME y bases de datos",
)

COURSE_EN = {
    "Análisis de la biodiversidad de especies y hábitats con QGIS, Google Earth Engine y OpenData": "Biodiversity analysis with QGIS, Google Earth Engine and OpenData",
    "Digitalización: Agricultura 4.0": "Digitalisation: Agriculture 4.0",
    "Python avanzado": "Advanced Python",
    "Scripts de geoprocesamiento con Python en ArcGIS Pro": "Geoprocessing scripts with Python in ArcGIS Pro",
    "Machine Learning para Investigación Científica - Random Forest, Boosting y Técnicas Avanzadas de Interpretación": "Machine learning for scientific research: random forests, boosting and advanced interpretation",
    "Deep learning con KERAS para investigación científica": "Deep learning with Keras for scientific research",
    "Desarrollo web con ArcGIS API para JavaScript": "Web development with the ArcGIS API for JavaScript",
    "Servicios sobre Docker": "Services with Docker",
    "Adapted numerical methods for Differential Equations": "Adapted numerical methods for differential equations",
    "FME y bases de datos": "FME and databases",
}

TOOLS = {
    "es": (
        "DiFLUsion — Sistema de alerta espacio-temporal para influenza aviar de alta patogenicidad",
        "ProtectIA — Herramienta geoespacial multicriterio para la prevención de la influenza aviar",
        "SMART-E — Surveillance and Monitoring of Antimicrobial Resistance in the Environment",
        "WiBISS — Wild Boar Immunization Strategy Simulator",
        "ASF FrontWave — African Swine Fever Front-Wave Rapid Risk Assessment",
        "Dashboards — Paneles epidemiológicos de vigilancia y evaluación del riesgo",
    ),
    "en": (
        "DiFLUsion — Spatio-temporal warning system for highly pathogenic avian influenza",
        "ProtectIA — Multi-criteria geospatial tool for avian-influenza prevention",
        "SMART-E — Surveillance and Monitoring of Antimicrobial Resistance in the Environment",
        "WiBISS — Wild Boar Immunization Strategy Simulator",
        "ASF FrontWave — African Swine Fever Front-Wave Rapid Risk Assessment",
        "Dashboards — Epidemiological dashboards for surveillance and risk assessment",
    ),
}

SKILLS = {
    "es": (
        "Python, MATLAB, Mathematica y Fortran",
        "QGIS, ArcGIS Pro/Online, Google Earth Engine y teledetección",
        "Modelización matemática, análisis espacio-temporal y aprendizaje automático",
        "Docker, desarrollo de aplicaciones geoespaciales y visualización de datos",
        "LaTeX y Microsoft Office",
        "Inglés profesional",
    ),
    "en": (
        "Python, MATLAB, Mathematica and Fortran",
        "QGIS, ArcGIS Pro/Online, Google Earth Engine and remote sensing",
        "Mathematical modelling, spatio-temporal analysis and machine learning",
        "Docker, geospatial application development and data visualisation",
        "LaTeX and Microsoft Office",
        "Professional English",
    ),
}


def text(value: Any) -> str:
    if value is None:
        return ""
    # En el PDF usamos guiones ASCII para asegurar una salida tipográfica
    # uniforme en lectores, impresoras y sistemas de selección de personal.
    normalized = re.sub(r"\s*[\u2010-\u2015]\s*", " - ", str(value))
    return re.sub(r"[ \t]+", " ", normalized).strip()


def esc(value: Any) -> str:
    return html.escape(text(value), quote=True)


def year(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return str(value.year)
    match = re.search(r"\b(19|20)\d{2}\b", text(value))
    return match.group(0) if match else ""


def fmt_hours(value: float) -> str:
    return f"{value:.1f}".replace(".", ",") if value % 1 else str(int(value))


def translate(value: Any, language: str) -> str:
    raw = text(value)
    return translate_value(raw) if language == "en" else raw


def is_pdf_record(record: dict[str, Any]) -> bool:
    return text(record.get("CV PDF")).upper() == "SI"


def load_records(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Actividades"]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    headers = [text(value) for value in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:] if any(value is not None for value in row)]


class TimelineParser(HTMLParser):
    """Extrae el timeline manual sin introducir una dependencia HTML adicional."""

    FIELDS = {"tl-period": "period", "tl-role": "role", "tl-org": "org", "tl-desc": "description", "tags": "tags"}

    def __init__(self) -> None:
        super().__init__()
        self.depth = 0
        self.current: dict[str, Any] | None = None
        self.root_depth: int | None = None
        self.active_fields: list[tuple[int, str]] = []
        self.entries: list[dict[str, str]] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag != "div":
            return
        self.depth += 1
        attr = dict(attrs)
        classes = set((attr.get("class") or "").split())
        root_kind = "phd" if "phd-floater" in classes else "timeline" if "tl-item" in classes else ""
        if root_kind:
            self.current = {"kind": root_kind}
            self.root_depth = self.depth
            self.active_fields = []
        if self.current:
            for class_name, field in self.FIELDS.items():
                if class_name in classes:
                    self.current.setdefault(field, [])
                    self.active_fields.append((self.depth, field))

    def handle_data(self, data: str) -> None:
        if self.current and self.active_fields and data.strip():
            self.current[self.active_fields[-1][1]].append(data.strip())

    def handle_endtag(self, tag: str) -> None:
        if tag != "div":
            return
        self.active_fields = [(depth, field) for depth, field in self.active_fields if depth != self.depth]
        if self.current and self.root_depth == self.depth:
            cleaned = {"kind": text(self.current.get("kind"))}
            for key in ("period", "role", "org", "description", "tags"):
                cleaned[key] = re.sub(r"\s+", " ", " ".join(self.current.get(key, []))).strip()
            self.entries.append(cleaned)
            self.current = None
            self.root_depth = None
            self.active_fields = []
        self.depth -= 1


def parse_timeline(path: Path) -> list[dict[str, str]]:
    parser = TimelineParser()
    parser.feed(path.read_text(encoding="utf-8"))
    return parser.entries


def register_fonts() -> None:
    fonts = {
        "CV-Regular": Path(r"C:\Windows\Fonts\arial.ttf"),
        "CV-Bold": Path(r"C:\Windows\Fonts\arialbd.ttf"),
        "CV-Italic": Path(r"C:\Windows\Fonts\ariali.ttf"),
    }
    for name, path in fonts.items():
        if path.is_file():
            pdfmetrics.registerFont(TTFont(name, str(path)))
    if "CV-Regular" not in pdfmetrics.getRegisteredFontNames():
        raise FileNotFoundError("No se encontraron las fuentes Arial necesarias para generar el PDF.")


def build_styles() -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    return {
        "name": ParagraphStyle("CVName", parent=base["Normal"], fontName="CV-Bold", fontSize=24, leading=27, textColor=DARK, spaceAfter=2),
        "title": ParagraphStyle("CVTitle", parent=base["Normal"], fontName="CV-Bold", fontSize=10.5, leading=13, textColor=GREEN),
        "contact": ParagraphStyle("CVContact", parent=base["Normal"], fontName="CV-Regular", fontSize=7.4, leading=10, textColor=MUTED),
        "section": ParagraphStyle("CVSection", parent=base["Normal"], fontName="CV-Bold", fontSize=10.5, leading=13, textColor=GREEN, spaceBefore=6, spaceAfter=3, keepWithNext=True),
        "body": ParagraphStyle("CVBody", parent=base["Normal"], fontName="CV-Regular", fontSize=8.25, leading=10.5, textColor=DARK, spaceAfter=2),
        "small": ParagraphStyle("CVSmall", parent=base["Normal"], fontName="CV-Regular", fontSize=7.4, leading=9.4, textColor=MUTED),
        "entry_title": ParagraphStyle("CVEntryTitle", parent=base["Normal"], fontName="CV-Bold", fontSize=8.5, leading=10.6, textColor=DARK),
        "entry_meta": ParagraphStyle("CVEntryMeta", parent=base["Normal"], fontName="CV-Regular", fontSize=7.4, leading=9.2, textColor=MUTED, spaceBefore=1),
        "period": ParagraphStyle("CVPeriod", parent=base["Normal"], fontName="CV-Bold", fontSize=7.5, leading=9.4, textColor=CLAY, alignment=TA_RIGHT),
        "citation": ParagraphStyle("CVCitation", parent=base["Normal"], fontName="CV-Regular", fontSize=7.15, leading=9.1, textColor=DARK, leftIndent=10, firstLineIndent=-10, spaceAfter=3),
        "bullet": ParagraphStyle("CVBullet", parent=base["Normal"], fontName="CV-Regular", fontSize=7.7, leading=9.7, textColor=DARK, leftIndent=9, firstLineIndent=-7, bulletIndent=0, spaceAfter=1.5),
    }


def section_header(title: str, styles: dict[str, ParagraphStyle]) -> list[Any]:
    return [Spacer(1, 2.5 * mm), Paragraph(esc(title), styles["section"]), HRFlowable(width="100%", thickness=0.65, color=LINE, spaceAfter=3.2 * mm)]


def entry_flowable(
    entry: dict[str, str],
    styles: dict[str, ParagraphStyle],
    *,
    description: bool = False,
    language: str = "es",
) -> KeepTogether:
    body = [Paragraph(esc(entry.get("role")), styles["entry_title"])]
    if entry.get("org"):
        body.append(Paragraph(esc(entry["org"]), styles["entry_meta"]))
    if description and entry.get("description"):
        body.append(Paragraph(esc(entry["description"]), styles["body"]))
    table = Table(
        [[Paragraph(esc(entry.get("period")).replace("actualidad", "present") if language == "en" else esc(entry.get("period")), styles["period"]), body]],
        colWidths=[28 * mm, 151 * mm],
        hAlign="LEFT",
        spaceBefore=1.5 * mm,
        spaceAfter=1.6 * mm,
    )
    table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (0, 0), 5 * mm),
        ("RIGHTPADDING", (1, 0), (1, 0), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    return KeepTogether([table])


def bold_candidate(authors: str) -> str:
    rendered = esc(authors)
    variants = (
        "Pablo Ibáñez-Porras", "Pablo Ibañez-Porras", "Pablo Ibañez", "Ibañez Pablo",
        "Ibáñez-Porras P", "Ibañez-Porras P", "Ibáñez-Porras, P.", "Ibañez-Porras, P.",
        "Ibáñez Porras P", "Ibañez Porras P", "Ibáñez Porras", "Ibañez Porras",
    )
    for variant in sorted(variants, key=len, reverse=True):
        safe = esc(variant)
        rendered = rendered.replace(safe, f"<b>{safe}</b>")
    return rendered


def citation(record: dict[str, Any], language: str) -> str:
    authors = bold_candidate(text(record.get("Autores")) or "P. Ibáñez-Porras")
    published = year(record.get("Fecha inicio")) or "s. f."
    title = esc(record.get("Nombre"))
    venue = esc(record.get("Revista") or record.get("Organizado"))
    doi = text(record.get("DOI"))
    link = text(record.get("Enlace"))
    tail = f" <i>{venue}</i>." if venue else ""
    if doi:
        url = f"https://doi.org/{re.sub(r'^https?://(?:dx\.)?doi\.org/', '', doi, flags=re.I)}"
        tail += f' <link href="{esc(url)}" color="#2F8F52">{esc(url)}</link>'
    elif link:
        tail += f' <link href="{esc(link)}" color="#2F8F52">{"Link" if language == "en" else "Enlace"}</link>'
    return f"{authors} ({published}). {title}.{tail}"


def bullet(text_value: str, styles: dict[str, ParagraphStyle]) -> Paragraph:
    return Paragraph(f"•&nbsp;&nbsp;{esc(text_value)}", styles["bullet"])


def footer(canvas: Any, doc: Any, language: str) -> None:
    canvas.saveState()
    canvas.setStrokeColor(LINE)
    canvas.setLineWidth(0.5)
    canvas.line(16 * mm, 11 * mm, A4[0] - 16 * mm, 11 * mm)
    canvas.setFont("CV-Regular", 7)
    canvas.setFillColor(MUTED)
    canvas.drawString(16 * mm, 7.2 * mm, "Pablo Ibáñez-Porras")
    canvas.drawRightString(A4[0] - 16 * mm, 7.2 * mm, f"{LABELS[language]['page']} {doc.page}")
    canvas.restoreState()


def header_story(language: str, styles: dict[str, ParagraphStyle]) -> list[Any]:
    email = '<link href="mailto:pabloibanezporras@gmail.com" color="#2F8F52">pabloibanezporras@gmail.com</link>'
    linkedin = '<link href="https://www.linkedin.com/in/pablo-ip" color="#2F8F52">LinkedIn</link>'
    orcid = '<link href="https://orcid.org/0009-0004-0618-8646" color="#2F8F52">ORCID</link>'
    contacts = f"Madrid, España · +34 916 202 300 ext. 2185 · {email} · {linkedin} · {orcid}"
    if language == "en":
        contacts = contacts.replace("España", "Spain")
    block = Table([
        [Paragraph("Pablo Ibáñez-Porras", styles["name"])],
        [Paragraph(LABELS[language]["title"], styles["title"])],
        [Paragraph(contacts, styles["contact"])],
    ], colWidths=[179 * mm])
    block.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), PALE),
        ("BOX", (0, 0), (-1, -1), 0.8, LINE),
        ("LEFTPADDING", (0, 0), (-1, -1), 7 * mm),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7 * mm),
        ("TOPPADDING", (0, 0), (-1, 0), 5 * mm),
        ("BOTTOMPADDING", (0, 2), (-1, 2), 5 * mm),
        ("TOPPADDING", (0, 1), (-1, 2), 1.2 * mm),
        ("BOTTOMPADDING", (0, 0), (-1, 1), 1.2 * mm),
    ]))
    return [block]


def generate_pdf(
    html_path: Path,
    pdf_path: Path,
    browser: Path | None = None,
    *,
    records: list[dict[str, Any]] | None = None,
    language: str | None = None,
    excel_path: Path | None = None,
) -> None:
    del browser  # Conservado para compatibilidad con llamadas anteriores.
    html_path = html_path.resolve()
    pdf_path = pdf_path.resolve()
    if not html_path.is_file():
        raise FileNotFoundError(f"No se encuentra el HTML: {html_path}")
    language = language or ("en" if "-en" in html_path.stem.lower() else "es")
    if language not in LABELS:
        raise ValueError(f"Idioma no compatible: {language}")
    records = records if records is not None else load_records((excel_path or DEFAULT_EXCEL).resolve())
    pdf_records = [record for record in records if is_pdf_record(record)]
    timeline = parse_timeline(html_path)
    register_fonts()
    styles = build_styles()

    story: list[Any] = header_story(language, styles)
    story.extend(section_header(LABELS[language]["profile"], styles))
    story.append(Paragraph(esc(LABELS[language]["profile_text"]), styles["body"]))

    contracted_tokens = {"contratado", "research appointment"}
    projects = [entry for entry in timeline if any(token in entry.get("tags", "").lower() for token in contracted_tokens)]
    story.extend(section_header(LABELS[language]["experience"], styles))
    for item in projects:
        story.append(entry_flowable(item, styles, description=True, language=language))

    education_tokens = {"formación académica", "academic education", "prácticas de grado", "undergraduate internship"}
    education = [entry for entry in timeline if entry.get("kind") == "phd" or any(token in entry.get("tags", "").lower() for token in education_tokens)]
    story.extend(section_header(LABELS[language]["education"], styles))
    for item in education:
        story.append(entry_flowable(item, styles, description=False, language=language))

    publications = [record for record in pdf_records if text(record.get("Tipo")) == "Publicación"]
    technical_venues = {"esrinews", "avinews"}
    scientific = [record for record in publications if text(record.get("Revista")).lower() not in technical_venues]
    technical = [record for record in publications if record not in scientific]
    scientific.sort(key=lambda item: item.get("Fecha inicio") or datetime.min, reverse=True)
    technical.sort(key=lambda item: item.get("Fecha inicio") or datetime.min, reverse=True)

    story.extend(section_header(LABELS[language]["publications"], styles))
    for record in scientific:
        story.append(KeepTogether([Paragraph(citation(record, language), styles["citation"])]))
    story.extend(section_header(LABELS[language]["technical"], styles))
    for record in technical:
        story.append(KeepTogether([Paragraph(citation(record, language), styles["citation"])]))

    forum_types = {"Congreso", "Conferencia", "Seminario", "Jornada", "Taller", "Webinar", "Asamblea General"}
    forums = [record for record in pdf_records if text(record.get("Tipo")) in forum_types]
    lead = [
        record for record in pdf_records
        if text(record.get("Tipo")) in {"Congreso", "Conferencia", "Asamblea General"}
        and text(record.get("Rol")) == "Principal"
    ]
    lead.sort(key=lambda item: item.get("Fecha inicio") or datetime.min, reverse=True)
    international = [record for record in forums if text(record.get("Internacional")).lower() in {"sí", "si", "internacional", "yes", "true"}]
    story.extend(section_header(LABELS[language]["activity"], styles))
    story.append(Paragraph(
        esc(LABELS[language]["activity_summary"].format(
            forums=len({text(item.get("Nombre Congreso / Detalle")) or text(item.get("Nombre")) for item in forums}),
            lead=len([item for item in forums if text(item.get("Rol")) in {"Principal", "Ponente"}]),
            international=len(international),
        )),
        styles["body"],
    ))
    story.append(Paragraph(esc(LABELS[language]["lead_contributions"]), styles["entry_title"]))
    for record in lead:
        event = text(record.get("Nombre Congreso / Detalle")) or text(record.get("Organizado"))
        kind = text(record.get("Tipo de ponencia")) or text(record.get("Tipo"))
        item = f"{translate(record.get('Nombre'), language)} — {kind}; {translate(event, language)} ({year(record.get('Fecha inicio'))})"
        story.append(bullet(item, styles))

    teaching = [record for record in pdf_records if text(record.get("Tipo")) == "Docencia" or text(record.get("Rol")) == "Docente"]
    seen_teaching: set[str] = set()
    story.extend(section_header(LABELS[language]["teaching"], styles))
    for record in teaching:
        name = text(record.get("Nombre"))
        if name in seen_teaching:
            continue
        seen_teaching.add(name)
        context = text(record.get("Nombre Congreso / Detalle")) or text(record.get("Revista")) or text(record.get("Organizado"))
        item = f"{translate(name, language)} - {translate(context, language)} ({year(record.get('Fecha inicio'))})"
        story.append(bullet(item, styles))
    advisory = [record for record in pdf_records if text(record.get("Tipo")) == "Asesoría"]
    if advisory:
        advice_years = sorted({year(record.get("Fecha inicio")) for record in advisory if year(record.get("Fecha inicio"))})
        period = f" ({advice_years[0]}-{advice_years[-1]})" if advice_years else ""
        detailed = next((record for record in advisory if text(record.get("Detalle"))), None)
        advice_name = translate(detailed.get("Nombre"), language) if detailed else LABELS[language]["advice"]
        story.append(bullet(f"{advice_name}{period}.", styles))
    else:
        story.append(bullet(LABELS[language]["advice"], styles))

    training = [
        record for record in pdf_records
        if (text(record.get("Tipo")) == "Curso" and text(record.get("Rol")) != "Docente") or text(record.get("Rol")) == "Asistente"
    ]
    total_hours = sum(float(item["Horas"]) for item in training if isinstance(item.get("Horas"), (int, float)))
    by_name = {text(record.get("Nombre")): record for record in training}
    story.extend(section_header(LABELS[language]["courses"], styles))
    story.append(Paragraph(esc(LABELS[language]["courses_summary"].format(hours=fmt_hours(total_hours))), styles["body"]))
    for name in COURSE_SELECTION:
        record = by_name.get(name)
        if not record:
            continue
        shown = COURSE_EN.get(name, name) if language == "en" else name
        hours = record.get("Horas")
        organizer = text(record.get("Revista")) or text(record.get("Organizado"))
        meta = " · ".join(part for part in (f"{fmt_hours(float(hours))} h" if isinstance(hours, (int, float)) else "", translate(organizer, language)) if part)
        story.append(bullet(f"{shown} — {meta}" if meta else shown, styles))

    story.extend(section_header(LABELS[language]["tools"], styles))
    for item in TOOLS[language]:
        story.append(bullet(item, styles))

    outreach = [record for record in pdf_records if text(record.get("Tipo")) == "Divulgación"]
    seen_outreach: set[tuple[str, str]] = set()
    story.extend(section_header(LABELS[language]["outreach"], styles))
    for record in outreach:
        key = (text(record.get("Nombre")), text(record.get("Nombre Congreso / Detalle")))
        if key in seen_outreach:
            continue
        seen_outreach.add(key)
        item = f"{translate(record.get('Nombre'), language)} — {translate(record.get('Nombre Congreso / Detalle'), language)} ({year(record.get('Fecha inicio'))})"
        story.append(bullet(item, styles))

    recognition = [record for record in pdf_records if text(record.get("Tipo")) in {"Premio", "Beca", "Estancia"}]
    story.extend(section_header(LABELS[language]["recognition"], styles))
    for record in recognition:
        institution = text(record.get("Revista")) or text(record.get("Organizado"))
        item = " · ".join(part for part in (translate(record.get("Nombre"), language), translate(institution, language), year(record.get("Fecha inicio"))) if part)
        story.append(bullet(item, styles))

    story.extend(section_header(LABELS[language]["skills"], styles))
    for item in SKILLS[language]:
        story.append(bullet(item, styles))

    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    handle, temporary_name = tempfile.mkstemp(prefix=".cv-documental-", suffix=".pdf", dir=pdf_path.parent)
    os.close(handle)
    temporary_pdf = Path(temporary_name)
    temporary_pdf.unlink(missing_ok=True)
    try:
        document = SimpleDocTemplate(
            str(temporary_pdf),
            pagesize=A4,
            rightMargin=16 * mm,
            leftMargin=16 * mm,
            topMargin=14 * mm,
            bottomMargin=16 * mm,
            title=f"CV - Pablo Ibáñez-Porras ({language.upper()})",
            author="Pablo Ibáñez-Porras",
            subject="Curriculum vitae",
        )
        document.build(
            story,
            onFirstPage=lambda canvas, doc: footer(canvas, doc, language),
            onLaterPages=lambda canvas, doc: footer(canvas, doc, language),
        )
        validate_pdf(temporary_pdf)
        os.replace(temporary_pdf, pdf_path)
    finally:
        temporary_pdf.unlink(missing_ok=True)
    validate_pdf(pdf_path)


def validate_pdf(path: Path) -> None:
    if not path.is_file() or path.stat().st_size < 10_000:
        raise RuntimeError(f"El PDF no se generó correctamente: {path}")
    with path.open("rb") as stream:
        if stream.read(5) != b"%PDF-":
            raise RuntimeError(f"El archivo generado no es un PDF válido: {path}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--spanish-html", type=Path, default=CV_DIR / "index.html")
    parser.add_argument("--english-html", type=Path, default=CV_DIR / "index-en.html")
    parser.add_argument("--spanish-pdf", type=Path, default=CV_DIR / "CV Pablo Ibáñez-Porras.pdf")
    parser.add_argument("--english-pdf", type=Path, default=CV_DIR / "CV Pablo Ibáñez-Porras EN.pdf")
    parser.add_argument("--browser", type=Path, help=argparse.SUPPRESS)
    args = parser.parse_args()

    records = load_records(args.excel.resolve())
    generate_pdf(args.spanish_html, args.spanish_pdf, records=records, language="es")
    print(f"PDF documental generado: {args.spanish_pdf.resolve()}")
    generate_pdf(args.english_html, args.english_pdf, records=records, language="en")
    print(f"English document PDF generated: {args.english_pdf.resolve()}")


if __name__ == "__main__":
    main()
