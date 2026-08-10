#!/usr/bin/env python3
"""Regenera las secciones automáticas del CV a partir de Actividades.xlsx.

El timeline de formación y experiencia queda fuera de los bloques AUTO y nunca se
reescribe. El contenido se inserta entre marcadores HTML explícitos.
"""

from __future__ import annotations

import argparse
import html
import os
import re
import tempfile
import unicodedata
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from generar_pdf import generate_pdf
from traducir_cv import write_english


CV_DIR = Path(__file__).resolve().parents[1]
EXCEL_DIR = CV_DIR.parent / "Certificados Académicos"
EXCEL_CANDIDATES = [EXCEL_DIR / "Actividades.xlsx", EXCEL_DIR / "Actividades.actualizado.xlsx"]
DEFAULT_EXCEL = max(
    (candidate for candidate in EXCEL_CANDIDATES if candidate.exists()),
    key=lambda candidate: candidate.stat().st_mtime,
    default=EXCEL_CANDIDATES[0],
)
DEFAULT_HTML = CV_DIR / "index.html"

MONTHS = (
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
)

AREA_LABELS = {
    "gis": "GIS y teledetección",
    "data": "IA, datos y programación",
    "maths": "Matemáticas y modelización",
    "one_health": "One Health y agrotecnología",
    "transversal": "Investigación y competencias",
}

AREA_FALLBACK = {
    "gis": "ArcGIS Pro y Online, QGIS, Google Earth Engine, teledetección y desarrollo de aplicaciones geoespaciales.",
    "data": "Aprendizaje automático y deep learning aplicados a investigación, Python, Docker, análisis estadístico y flujos reproducibles.",
    "maths": "Métodos numéricos, ecuaciones diferenciales y simulación científica aplicada.",
    "one_health": "Análisis de riesgo sanitario, enfermedades transfronterizas, patología veterinaria y digitalización del sector agropecuario.",
    "transversal": "Comunicación científica, revisión bibliográfica, investigación cualitativa, idiomas y competencias profesionales.",
}

HIGHLIGHT_DETAILS = {
    "smart-e": "Marco espacial para estandarizar la vigilancia ambiental de la resistencia antimicrobiana. Integra fuentes antropogénicas, contexto territorial y puntos de muestreo para facilitar comparaciones reproducibles entre territorios.",
    "from science to policy": "Evolución de DiFLUsion hacia un sistema operativo de alerta temprana desarrollado con los responsables de vigilancia. La contribución conecta modelización espacio-temporal, necesidades de usuario y transferencia a política sanitaria.",
    "frontwave": "Visor de evaluación rápida que interpreta la velocidad y dirección del frente epidémico de peste porcina africana. Facilita la priorización territorial de la vigilancia y de las medidas de control.",
    "fuentes de emisión": "Cartografía de fuentes antropogénicas asociadas a resistencia antimicrobiana para caracterizar su presión ambiental. El análisis integra información espacial heterogénea en una lectura territorial comparable.",
    "protectia": "Herramienta geoespacial multicriterio que combina riesgo epidemiológico, bioseguridad y contexto territorial. Permite priorizar comarcas y explotaciones avícolas para apoyar la prevención de la influenza aviar.",
    "geoprocesamiento": "Aplicación de flujos de geoprocesamiento a la vigilancia y evaluación rápida de enfermedades emergentes. La propuesta traduce datos epidemiológicos y ambientales en indicadores espaciales útiles para la decisión.",
    "wibiss": "Simulación espacial de estrategias de vacunación del jabalí frente a la peste porcina africana. Los escenarios se traducen en restricciones comerciales y pérdidas potencialmente evitadas para los productores porcinos.",
    "diflusion: a tool": "Presentación del sistema DiFLUsion para la vigilancia de influenza aviar mediante alertas semanales. Combina información de brotes, condiciones ambientales y movimientos de aves silvestres.",
    "epidemiología y análisis espacial": "Recorrido aplicado por el uso de SIG, análisis espacio-temporal y visualización para comprender dinámicas de enfermedad. El enfoque One Health conecta salud animal, salud pública y medio ambiente.",
    "contaminación de suelos": "Análisis espacial de la presencia de antibióticos y genes de resistencia en suelos. La contribución aborda la relación entre fuentes de presión, medio ambiente y vigilancia One Health.",
    "marinestrandingviewer": "Herramienta web para registrar y visualizar varamientos de mamíferos marinos. Centraliza observaciones georreferenciadas y facilita su exploración para investigación y vigilancia.",
    "jornadas vigilancia": "Presentación de DiFLUsion como sistema de alerta en tiempo real para influenza aviar. El trabajo muestra cómo convertir datos de vigilancia en mapas de riesgo de actualización periódica.",
    "sistema de alerta a tiempo real": "Presentación de DiFLUsion como sistema de alerta en tiempo real para influenza aviar. El trabajo muestra cómo convertir datos de vigilancia en mapas de riesgo de actualización periódica.",
    "hormigas, gallinas": "Comunicación divulgativa sobre el uso de modelos matemáticos y datos espaciales para entender epidemias. Conecta conceptos de redes, movilidad y transmisión con ejemplos del ámbito animal.",
    "ciencia ciudadana": "Aplicación directa de observaciones de SEO/BirdLife al sistema DiFLUsion. La contribución muestra el valor de la ciencia ciudadana para reforzar la vigilancia de influenza aviar.",
}

LEAD_HIGHLIGHT_DETAILS = {
    "smart-e": "Marco espacial que integra fuentes antropogénicas, contexto territorial y puntos de muestreo para estandarizar la vigilancia ambiental de la resistencia antimicrobiana y facilitar comparaciones reproducibles entre territorios.",
    "from science to policy": "DiFLUsion evoluciona hacia un sistema operativo de alerta temprana codesarrollado con responsables de vigilancia que conecta modelización espacio-temporal, necesidades de usuario y política sanitaria.",
    "frontwave": "Visor de evaluación rápida que interpreta la velocidad y dirección del frente epidémico de peste porcina africana para priorizar territorialmente la vigilancia y las medidas de control.",
    "wibiss": "Simulación espacial de estrategias de vacunación del jabalí frente a la peste porcina africana que traduce los escenarios en restricciones comerciales y pérdidas potencialmente evitadas para los productores porcinos.",
    "marinestrandingviewer": "Herramienta web que centraliza observaciones georreferenciadas para registrar, visualizar y explorar varamientos de mamíferos marinos con fines de investigación y vigilancia.",
}

PUBLICATION_DETAILS = {
    "wibiss": "Aplicación de la herramienta WiBISS al norte de Italia para explorar cómo distintas estrategias de vacunación del jabalí podrían contribuir al control de la peste porcina africana y a reducir sus consecuencias económicas.",
    "diflusion": "Aplicación de la herramienta DiFLUsion en España para evaluar el riesgo de introducción de IAAP asociado a los movimientos migratorios de aves.",
}

ADVISORY_NAME = "Asesoría técnica sobre la situación de la Influenza Aviar de Alta Patogenicidad (IAAP) y el uso de DiFLUsion como sistema de alerta temprana"
ADVISORY_DETAIL = "Múltiples sesiones de colaboración con organismos nacionales e internacionales, administraciones y otros actores implicados en la vigilancia sanitaria, aportando análisis epidemiológico, interpretación del riesgo y herramientas para apoyar la toma de decisiones y la preparación frente a posibles introducciones de la enfermedad."

FEATURED_TEACHING = {
    "the overlooked role of passeriformes in avian influenza and antimicrobial resistance amid climate change",
    "training course on risk analysis for transboundary animal disease control",
}
FEATURED_OUTREACH = {"detectives de epidemias: gallinas y matematicas"}

TEACHING_FALLBACK = {
    "risk analysis": "Curso internacional sobre análisis de riesgo para el control de enfermedades animales transfronterizas, impartido para la representación subregional de WOAH en el Sudeste Asiático.",
    "passeriformes": "Taller docente sobre el papel de los paseriformes en la conectividad de influenza aviar y resistencia antimicrobiana bajo escenarios de cambio climático.",
    "implementación de la ia": "Formación aplicada sobre inteligencia artificial, big data y digitalización en investigación de sanidad animal y medicamentos veterinarios.",
    "científic@s en prácticas": "Tutorización de estudiantes durante una experiencia de inmersión en investigación científica en el CISA-INIA/CSIC.",
    "distintas enfermedades": "Supervisión de prácticas externas centradas en el análisis de la distribución espacio-temporal de enfermedades animales.",
    "suavizado temporal": "Supervisión de un Trabajo de Fin de Grado sobre herramientas matemáticas para el suavizado temporal de rutas migratorias.",
    "brotes de iaap": "Supervisión de prácticas externas sobre el análisis de la distribución espacio-temporal de brotes de influenza aviar de alta patogenicidad en España.",
}

OUTREACH_FALLBACK = {
    "detectives de epidemias": "Charla sobre cómo la física, las matemáticas y los datos ayudan a seguir y anticipar una epidemia.",
    "científic@s en prácticas": "Tutorización de estudiantes para conocer de primera mano el trabajo y los métodos de la investigación científica.",
    "auténtico detective": "Taller participativo centrado en observación, datos y toma de decisiones frente a brotes.",
    "epidemias y pandemias": "Actividad participativa para comprender de forma lúdica la propagación y las cadenas de transmisión de agentes infecciosos.",
}


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def plain(value: Any) -> str:
    return unicodedata.normalize("NFD", text(value)).encode("ascii", "ignore").decode().lower()


def esc(value: Any) -> str:
    return html.escape(text(value), quote=True)


def year(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return str(value.year)
    match = re.search(r"\b(19|20)\d{2}\b", text(value))
    return match.group(0) if match else ""


def month_year(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return f"{MONTHS[value.month - 1]} {value.year}"
    return year(value)


def fmt_hours(value: float) -> str:
    return f"{value:.1f}".replace(".", ",") if value % 1 else str(int(value))


def detail_for(record: dict[str, Any], fallbacks: dict[str, str]) -> str:
    explicit = text(record.get("Detalle"))
    if explicit:
        return explicit
    name = plain(record.get("Nombre"))
    for key, value in fallbacks.items():
        if plain(key) in name:
            return value
    return "Participación vinculada a investigación, transferencia de conocimiento y aplicación de métodos cuantitativos al ámbito One Health."


def doi_url(record: dict[str, Any]) -> str:
    link = text(record.get("Enlace"))
    if link:
        return link
    doi = text(record.get("DOI"))
    if not doi:
        return ""
    doi = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", doi, flags=re.I)
    return f"https://doi.org/{doi}"


def record_card(
    record: dict[str, Any],
    class_name: str,
    content: str,
    *,
    fallback_tag: str = "article",
) -> str:
    """Renderiza una tarjeta completa como enlace cuando el registro tiene URL o DOI."""
    url = doi_url(record)
    if url:
        label = f"Abrir enlace: {text(record.get('Nombre'))}"
        return (
            f'<a class="{class_name} linked-card reveal" href="{esc(url)}" rel="noopener" target="_blank" '
            f'aria-label="{esc(label)}">{content}<span class="card-link-label">Abrir enlace ↗</span></a>'
        )
    return f'<{fallback_tag} class="{class_name} reveal">{content}</{fallback_tag}>'


def training_area(record: dict[str, Any]) -> str:
    combined = plain(" ".join(text(record.get(key)) for key in ("Nombre", "Nombre Congreso / Detalle", "Organizado")))
    if re.search(r"arcgis|qgis|google earth|teledeteccion|maxent|geoanalytics|story maps|web appbuilder|fme|esri", combined):
        return "gis"
    if re.search(r"python|machine learning|deep learning|keras|docker|jasp|inteligencia artificial|redes neuronales|aihub|big data", combined):
        return "data"
    if re.search(r"differential equations|metodos numericos|ecuaciones diferenciales|laser-plasma", combined):
        return "maths"
    if re.search(r"peste porcina|rhdv|patologia veterinaria|risk analysis|agricultura 4\.0|animal disease|diagnostico de laboratorio", combined):
        return "one_health"
    return "transversal"


def replace_block(source: str, name: str, content: str) -> str:
    start = f"<!-- AUTO:{name}:START -->"
    end = f"<!-- AUTO:{name}:END -->"
    pattern = re.compile(re.escape(start) + r".*?" + re.escape(end), re.S)
    if not pattern.search(source):
        raise RuntimeError(f"No se encontraron los marcadores {start} / {end}")
    return pattern.sub(f"{start}\n{content.rstrip()}\n{end}", source, count=1)


def load_records(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Actividades"]
        values = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    headers = [text(value) for value in values[0]]
    return [dict(zip(headers, row)) for row in values[1:] if any(value is not None for value in row)]


def render_publications(records: list[dict[str, Any]]) -> str:
    publications = [r for r in records if text(r.get("Tipo")) == "Publicación"]
    technical_venues = {"esrinews", "avinews"}
    scientific = [r for r in publications if plain(r.get("Revista")) not in technical_venues]
    technical = [r for r in publications if r not in scientific]
    principal = [r for r in scientific if text(r.get("Rol")) == "Principal"]
    secondary = [r for r in scientific if text(r.get("Rol")) != "Principal"]
    featured = []
    for record in principal:
        url = doi_url(record)
        journal = text(record.get("Revista"))
        citation = " · ".join(part for part in (journal, f"doi:{text(record.get('DOI'))}" if record.get("DOI") else "") if part)
        featured.append(
            f'<a class="publication-link reveal" href="{esc(url)}" rel="noopener" target="_blank">'
            f'<article class="publication-lead"><span class="lead-badge">Publicación · {year(record.get("Fecha inicio"))}</span>'
            f'<h3>{esc(record.get("Nombre"))}</h3><p>{esc(detail_for(record, PUBLICATION_DETAILS))}</p>'
            f'<div class="pub-citation">{esc(citation)} ↗</div></article></a>'
        )
    compact = []
    for record in secondary:
        url = doi_url(record)
        venue = text(record.get("Revista")) or text(record.get("Organizado")) or "Publicación"
        tag = "a" if url else "article"
        attrs = f' href="{esc(url)}" rel="noopener" target="_blank"' if url else ""
        compact.append(
            f'<{tag} class="compact-publication reveal"{attrs}><time>{year(record.get("Fecha inicio"))}</time>'
            f'<strong>{esc(record.get("Nombre"))}</strong><span>{esc(venue)}{" ↗" if url else ""}</span></{tag}>'
        )
    technical_cards = []
    for record in technical:
        url = doi_url(record)
        venue = text(record.get("Revista")) or "Medio profesional"
        tag = "a" if url else "article"
        attrs = f' href="{esc(url)}" rel="noopener" target="_blank"' if url else ""
        technical_cards.append(
            f'<{tag} class="technical-publication reveal"{attrs}><time>{year(record.get("Fecha inicio"))}</time>'
            f'<strong>{esc(record.get("Nombre"))}</strong><span>{esc(venue)}{" ↗" if url else ""}</span></{tag}>'
        )
    return (
        '<div class="publication-featured" id="projects">\n' + "\n".join(featured) + '\n</div>\n'
        '<div class="secondary-publications"><h3>Otras publicaciones científicas y contribuciones como coautor</h3>'
        '<div class="compact-publication-list">\n' + "\n".join(compact) + "\n</div></div>\n"
        '<div class="technical-publications"><h3>Artículos técnico-profesionales y de transferencia</h3>'
        '<div class="technical-publication-list">\n' + "\n".join(technical_cards) + "\n</div></div>"
    )


def render_training(records: list[dict[str, Any]]) -> tuple[str, float]:
    training = [
        r for r in records
        if (text(r.get("Tipo")) == "Curso" and text(r.get("Rol")) != "Docente")
        or text(r.get("Rol")) == "Asistente"
    ]
    grouped: dict[str, list[dict[str, Any]]] = {key: [] for key in AREA_LABELS}
    for record in training:
        grouped[training_area(record)].append(record)
    cards = []
    total = 0.0
    for key, label in AREA_LABELS.items():
        items = grouped[key]
        hours = sum(float(r["Horas"]) for r in items if isinstance(r.get("Horas"), (int, float)))
        total += hours
        explicit = next((text(r.get("Detalle")) for r in items if text(r.get("Detalle"))), AREA_FALLBACK[key])
        cards.append(
            f'<article class="training-summary-card reveal"><b>+{fmt_hours(hours)} h</b><h4>{esc(label)}</h4>'
            f'<p>{esc(explicit)}</p></article>'
        )
    return (
        '<div class="subsection-head reveal"><h3>Formación continua por especialidad</h3></div>\n'
        '<div class="training-summary-grid">\n' + "\n".join(cards) + "\n</div>", total
    )


def render_activity(records: list[dict[str, Any]]) -> str:
    forum_types = {"Congreso", "Conferencia", "Seminario", "Jornada", "Taller", "Webinar", "Asamblea General"}
    forums = [r for r in records if text(r.get("Tipo")) in forum_types]
    unique_events = {text(r.get("Nombre Congreso / Detalle")) or text(r.get("Nombre")) for r in forums}
    oral = sum(1 for r in forums if "oral" in plain(r.get("Tipo de ponencia")))
    posters = sum(1 for r in forums if "poster" in plain(r.get("Tipo de ponencia")))
    lead_or_speaker = sum(1 for r in forums if text(r.get("Rol")) in {"Principal", "Ponente"})
    coauthors = sum(1 for r in forums if text(r.get("Rol")).startswith("Coautor"))
    highlights = [
        r for r in records
        if text(r.get("Tipo")) in {"Congreso", "Conferencia", "Asamblea General"}
        and text(r.get("Rol")) == "Principal"
    ]
    international = [r for r in highlights if plain(r.get("Internacional")) in {"si", "internacional", "yes", "true"}]
    national = [r for r in highlights if r not in international]
    metrics = (
        '<div class="activity-metrics reveal" aria-label="Resumen de participación científica">'
        f'<div class="metric-card"><b>{len(unique_events)}</b><span>encuentros científicos</span></div>'
        f'<div class="metric-card"><b>{lead_or_speaker}</b><span>como autor principal o ponente</span></div>'
        f'<div class="metric-card"><b>{oral}</b><span>comunicaciones orales</span></div>'
        f'<div class="metric-card"><b>{posters}</b><span>pósteres</span></div>'
        f'<div class="metric-card"><b>{coauthors}</b><span>aportaciones como coautor</span></div></div>'
    )
    cloud = (
        '<div class="topic-cloud reveal" aria-label="Áreas de participación científica">'
        '<span>Vigilancia de influenza aviar</span><span>Resistencia antimicrobiana ambiental</span>'
        '<span>Peste porcina africana</span><span>One Health e interoperabilidad</span>'
        '<span>Epidemiología espacial y SIG</span><span>Fauna silvestre y salud ambiental</span></div>'
    )
    cards = []
    for record in international:
        event = text(record.get("Nombre Congreso / Detalle")) or text(record.get("Organizado"))
        meta = " · ".join(part for part in (text(record.get("Tipo de ponencia")) or text(record.get("Tipo")), event, month_year(record.get("Fecha inicio"))) if part)
        content = (
            f'<div class="talk-meta">{esc(meta)}</div><h4>{esc(record.get("Nombre"))}</h4>'
            f'<p>{esc(detail_for(record, LEAD_HIGHLIGHT_DETAILS))}</p>'
        )
        cards.append(record_card(record, "lead-talk", content))
    compact_cards = []
    for record in national:
        event = text(record.get("Nombre Congreso / Detalle")) or text(record.get("Organizado")) or text(record.get("Tipo"))
        meta = " · ".join(part for part in (text(record.get("Tipo de ponencia")) or text(record.get("Tipo")), event) if part)
        content = (
            f'<time>{year(record.get("Fecha inicio"))}</time><strong>{esc(record.get("Nombre"))}</strong>'
            f'<span>{esc(meta)}</span>'
        )
        compact_cards.append(record_card(record, "compact-talk", content))
    advisory = [r for r in records if text(r.get("Tipo")) == "Asesoría"]
    advice_html = ""
    if advisory:
        first = next((record for record in advisory if text(record.get("Detalle"))), advisory[0])
        advice_detail = text(first.get("Detalle"))
        advice_name = text(first.get("Nombre")) if advice_detail else ADVISORY_NAME
        advice_detail = advice_detail or ADVISORY_DETAIL
        content = (
            f'<strong>{esc(advice_name)}</strong>'
            f'<p>{esc(advice_detail)}</p>'
        )
        advice_html = (
            '<div class="subsection-head reveal"><h3>Asesoría científico-técnica</h3></div>'
            + record_card(first, "advisory-note", content, fallback_tag="div")
        )
    return (
        metrics + "\n" + cloud + '\n<div class="subsection-head reveal"><h3>Contribuciones destacadas como autor principal</h3></div>\n'
        '<div class="lead-talks">\n' + "\n".join(cards) + "\n</div>\n"
        '<div class="secondary-talks"><h4>Otras contribuciones como autor principal</h4>'
        '<div class="compact-talk-list">\n' + "\n".join(compact_cards) + "\n</div></div>\n" + advice_html
    )


def render_teaching(records: list[dict[str, Any]]) -> str:
    teaching = [r for r in records if text(r.get("Tipo")) == "Docencia" or text(r.get("Rol")) == "Docente"]
    cards = []
    for record in teaching:
        featured = plain(record.get("Nombre")) in FEATURED_TEACHING
        kind = text(record.get("Nombre Congreso / Detalle")) or text(record.get("Tipo"))
        role = text(record.get("Rol"))
        label = " · ".join(part for part in (kind, role if role and role != kind else "", year(record.get("Fecha inicio"))) if part)
        content = (
            ('<span class="featured-card-label">Docencia internacional destacada</span>' if featured else '') +
            f'<div class="teaching-type">{esc(label)}</div><h4>{esc(record.get("Nombre"))}</h4>'
            f'<p>{esc(detail_for(record, TEACHING_FALLBACK))}</p>'
        )
        class_name = "teaching-card teaching-card--featured" if featured else "teaching-card"
        cards.append(record_card(record, class_name, content))
    forum_types = {"Seminario", "Jornada", "Taller", "Webinar"}
    forums = [r for r in records if text(r.get("Tipo")) in forum_types and text(r.get("Rol")) != "Asistente"]
    rows = []
    for record in forums:
        context = text(record.get("Nombre Congreso / Detalle")) or text(record.get("Organizado")) or text(record.get("Tipo"))
        meta = " · ".join(part for part in (context, month_year(record.get("Fecha inicio"))) if part)
        content = (
            f'<span class="event-badge">{esc(record.get("Rol") or record.get("Tipo"))}</span>'
            f'<div><strong>{esc(record.get("Nombre"))}</strong><span>{esc(meta)}</span></div>'
        )
        rows.append(record_card(record, "event-row", content, fallback_tag="div"))
    return (
        '<div class="teaching-grid">\n' + "\n".join(cards) + "\n</div>\n"
        '<div class="subsection-head reveal"><h3>Seminarios, jornadas, talleres y otros foros</h3></div>\n'
        '<div class="event-list">\n' + "\n".join(rows) + "\n</div>"
    )


def render_outreach(records: list[dict[str, Any]]) -> str:
    cards = []
    for record in records:
        if text(record.get("Tipo")) != "Divulgación":
            continue
        featured = plain(record.get("Nombre")) in FEATURED_OUTREACH
        label = " · ".join(part for part in (text(record.get("Nombre Congreso / Detalle")) or "Divulgación", year(record.get("Fecha inicio"))) if part)
        link = text(record.get("Enlace"))
        content = (
            ('<span class="featured-card-label">Divulgación destacada</span>' if featured else '') +
            f'<div class="card-kind">{esc(label)}</div><h3>{esc(record.get("Nombre"))}</h3>'
            f'<p>{esc(detail_for(record, OUTREACH_FALLBACK))}</p>'
        )
        class_name = "standalone-card standalone-card--featured" if featured else "standalone-card"
        if link:
            cards.append(
                f'<a class="{class_name} standalone-card--link reveal" href="{esc(link)}" rel="noopener" target="_blank" '
                f'aria-label="Abrir {esc(record.get("Nombre"))}">{content}<span class="standalone-card__link">Abrir enlace ↗</span></a>'
            )
        else:
            cards.append(f'<article class="{class_name} reveal">{content}</article>')
    return '<div class="standalone-grid">\n' + "\n".join(cards) + "\n</div>"


def render_recognition(records: list[dict[str, Any]]) -> str:
    selected = [r for r in records if text(r.get("Tipo")) in {"Premio", "Beca", "Estancia"}]
    cards = []
    for record in selected:
        meta = " · ".join(part for part in (text(record.get("Organizado")), text(record.get("Lugar")), month_year(record.get("Fecha inicio"))) if part)
        content = (
            f'<div class="year">{year(record.get("Fecha inicio"))}</div><h3>{esc(record.get("Nombre"))}</h3>'
            f'<p>{esc(text(record.get("Detalle")) or meta)}</p>'
        )
        cards.append(record_card(record, "recognition-card", content))
    return '<div class="recognition-grid">\n' + "\n".join(cards) + "\n</div>"


def replace_stat(source: str, label: str, value: int | float) -> str:
    number = f"{float(value):g}"
    pattern = re.compile(rf'(<div class="stat"><b data-count=")\d+(?:\.\d+)?("[^>]*>0</b><span>{re.escape(label)}</span></div>)')
    return pattern.sub(rf"\g<1>{number}\g<2>", source, count=1)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="Libro de actividades")
    parser.add_argument("--html", type=Path, default=DEFAULT_HTML, help="HTML de entrada y salida")
    parser.add_argument("--output", type=Path, help="Salida alternativa; por defecto sustituye --html")
    parser.add_argument("--english-output", type=Path, default=CV_DIR / "index-en.html", help="Salida de la versión inglesa")
    parser.add_argument("--skip-english", action="store_true", help="No generar la versión inglesa")
    parser.add_argument("--spanish-pdf", type=Path, default=CV_DIR / "CV Pablo Ibáñez-Porras.pdf", help="PDF en español")
    parser.add_argument("--english-pdf", type=Path, default=CV_DIR / "CV Pablo Ibáñez-Porras EN.pdf", help="PDF en inglés")
    parser.add_argument("--skip-pdf", action="store_true", help="No generar los PDF")
    parser.add_argument("--browser", type=Path, help="Ruta opcional a Edge, Chrome o Chromium")
    args = parser.parse_args()

    records = load_records(args.excel.resolve())
    source = args.html.read_text(encoding="utf-8")
    manual_timeline = re.search(r"<!-- MANUAL:TIMELINE:START -->(.*?)<!-- MANUAL:TIMELINE:END -->", source, re.S)
    if not manual_timeline:
        raise RuntimeError("Faltan los marcadores que protegen el timeline manual")
    timeline_snapshot = manual_timeline.group(1)

    training_html, total_hours = render_training(records)
    blocks = {
        "TRAINING": training_html,
        "PUBLICATIONS": render_publications(records),
        "ACADEMIC": render_activity(records),
        "TEACHING": render_teaching(records),
        "OUTREACH": render_outreach(records),
        "RECOGNITION": render_recognition(records),
    }
    for name, content in blocks.items():
        source = replace_block(source, name, content)

    publications = sum(1 for r in records if text(r.get("Tipo")) == "Publicación")
    forum_types = {"Congreso", "Conferencia", "Seminario", "Jornada", "Taller", "Webinar", "Asamblea General"}
    forums = [r for r in records if text(r.get("Tipo")) in forum_types]
    source = replace_stat(source, "Publicaciones", publications)
    source = replace_stat(source, "Foros científicos", len({text(r.get("Nombre Congreso / Detalle")) or text(r.get("Nombre")) for r in forums}))
    source = replace_stat(source, "Horas acreditadas", total_hours)

    current_timeline = re.search(r"<!-- MANUAL:TIMELINE:START -->(.*?)<!-- MANUAL:TIMELINE:END -->", source, re.S)
    if not current_timeline or current_timeline.group(1) != timeline_snapshot:
        raise RuntimeError("La verificación de integridad del timeline manual ha fallado")

    output = (args.output or args.html).resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    if output == args.html.resolve():
        handle, temp_name = tempfile.mkstemp(prefix="index.", suffix=".tmp", dir=output.parent)
        os.close(handle)
        temp_path = Path(temp_name)
        try:
            temp_path.write_text(source, encoding="utf-8", newline="\n")
            os.replace(temp_path, output)
        finally:
            temp_path.unlink(missing_ok=True)
    else:
        output.write_text(source, encoding="utf-8", newline="\n")

    counts = Counter(text(r.get("Tipo")) for r in records)
    print(f"CV generado: {output}")
    print(f"Registros: {len(records)} · publicaciones: {counts['Publicación']} · horas: {fmt_hours(total_hours)}")
    print("Timeline manual: verificado y sin cambios")
    if not args.skip_english:
        missing = write_english(output, args.english_output.resolve())
        print(f"English CV generated: {args.english_output.resolve()}")
        if missing:
            print(f"Traducciones pendientes de revisión: {len(missing)}")
    if not args.skip_pdf:
        generate_pdf(output, args.spanish_pdf.resolve(), args.browser, records=records, language="es")
        print(f"PDF generado: {args.spanish_pdf.resolve()}")
        english_html = args.english_output.resolve()
        if english_html.is_file():
            generate_pdf(english_html, args.english_pdf.resolve(), args.browser, records=records, language="en")
            print(f"English PDF generated: {args.english_pdf.resolve()}")


if __name__ == "__main__":
    main()
